import os.path
import sys
import traceback

import you_get
import openpyxl
import requests
import time
from settings import picture_download_interval, record_type, record_source, record_keyword, record_author, record_title, \
    record_introduction, record_tags, record_caption, record_url_article, record_url_pic, record_url_local, log_path, \
    excel_path, down_video, requests_timeout

requests.DEFAULT_RETRIES = 5  # 增加重试连接次数

def log(message:str):
    with open(log_path, 'a') as f:
        f.write(time.ctime())
        f.write(message + "\r\n")
    # logging.debug(traceback.format_exc())

def getResModel(type:str, source:str, keyword:str, author:str,
                title:str, introduction:str, tags:list, caption:str,
                url_article:str, url_pic:str, url_local:str):
    res = {
        record_type: type,
        record_source: source,
        record_keyword: keyword,
        record_author: author,
        record_title: title,
        record_introduction: introduction,
        record_tags: tags,
        record_caption: caption,
        record_url_article: url_article,
        record_url_pic: url_pic,
        record_url_local: url_local,
    }
    return res

def download_pic(url:str, path:str):
    r = requests.get(url, timeout=requests_timeout)
    try:
        open(path, 'wb').write(r.content) # 将内容写入图片
    except:
        log(traceback.format_exc())
    del r
    time.sleep(picture_download_interval)  # 停一下，别过分了

def download_video(video_url: str, path:str, name:str):
    if not down_video:
        return
    # 使用sys.argv内置方法，可以在代码中输入CMD命令
    sys.argv = ['you-get', '-o', path, '-O', name, video_url, "--debug"]
    # 通过you-get模块，实现下载
    you_get.main()

# excel
class Excel:
    # 全部关键词
    keywordDict = {
        # "设备":[],
        # "软件":[],
        # "形式":[],
        # "类型":[]
    }
    # 防止重复导入网页记录
    allHttps = dict()
    # 统计tag分布
    allTags = dict()

    def __init__(self, path:str):
        self.path = path
        self.excel = openpyxl.load_workbook(path)
        sheetnames = self.excel.sheetnames

        self.table = self.excel[sheetnames[0]]
        self.__initHttps()

        self.tagTable = self.excel[sheetnames[2]]
        self.__initTags()

        self.keywordTable = self.excel[sheetnames[1]]
        self.__initKeyWords()

    # 获取实际最大行行号
    def __getMaxRow(self, sheet):
        i = sheet.max_row
        print("伪最大行数：" + str(i))
        real_max_row = 0
        while i > 0:
            # print(i)
            if i % 500 == 0:
                print("正在检查有效行：" + str(i))
            if sheet[i][0].value is None:
                i = i-1
            else:
                real_max_row = i
                break
        print("真最大行数：" + str(real_max_row))
        return real_max_row

    # 读取链接记录
    def __initHttps(self):
        self.nrows = self.__getMaxRow(self.table) # 获得行数
        # self.nrows = 1
        self.ncolumns = self.table.max_column # 获得列数
        for i in range(1, self.nrows+1):
            self.__recordHttps(self.table.cell(i,8).value)

    # 读取tag记录
    def __initTags(self):
        tagRows = self.__getMaxRow(self.tagTable)
        for i in range(1, tagRows + 1):
            self.allTags[self.tagTable.cell(i, 1).value] = self.tagTable.cell(i, 2).value

    # 读取关键词
    def __initKeyWords(self):
        rows = self.__getMaxRow(self.keywordTable)
        column = self.keywordTable.max_column
        for i in range(1, rows+1):
            temp = []
            for j in range(2, column+1):
                try:
                    for item in self.keywordTable.cell(i,j).value.split("，"):
                        temp.append(item)
                except:
                    # log()
                    pass
            self.keywordDict[self.keywordTable.cell(i,1).value] = temp

    # 查重
    def checkHttpRepeat(self, url:str):
        return url in self.allHttps.keys()

    # 统计新增的网页链接
    def __recordHttps(self, url:str):
        if url in self.allHttps.keys():
            self.allHttps[url] += 1
            # return False
        else:
            self.allHttps[url] = 1
            # return True

    # 统计新增的tag
    def __recordTags(self, tag:str):
        if tag in self.allTags.keys():
            self.allTags[tag] += 1
        else:
            self.allTags[tag] = 1

    # 因为tags为列表
    def appendTagsFromList(self, tags:list):
        for tag in tags:
            self.__recordTags(tag)

    # 保存并记录tags
    def __saveTags(self):
        i = 1
        for key in self.allTags.keys():
            self.tagTable.cell(i, 1).value = key
            self.tagTable.cell(i, 2).value = self.allTags[key]
            i+=1

    def appendRecord(self, resDict:dict):
        aim = [resDict[record_type], resDict[record_source], resDict[record_keyword], resDict[record_author],
               resDict[record_title], resDict[record_introduction], ",".join(resDict[record_tags]),
               resDict[record_caption], resDict[record_url_article],
               resDict[record_url_pic], resDict[record_url_local]]
        column = 1
        self.nrows += 1
        for value in aim:
            self.table.cell(self.nrows, column).value = value
            column += 1

    def saveExcel(self):
        self.__saveTags()
        self.excel.save(self.path)

excel = Excel(excel_path)
